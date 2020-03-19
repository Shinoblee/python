import openpyxl

# wb = openpyxl.Workbook()
wb = openpyxl.load_work("transactions.xlsx")
print(wb.sheetnames)

sheet = wb["Sheet1"]

wb.create_sheet("Sheet2", 0)

cell = sheet["a1"]
cell.value
cell.row
cell.column
cell.coordinate
cell = sheet.cell(row=1, coluimn=1)

for row in range(1, sheet.max_row + 1):
    for column in range(1, sheet.max_column + 1):
        cell = sheet.cell(row, column)
        print(cell.value)

column = sheet["a"]
